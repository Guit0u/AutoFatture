import pdfplumber
import openpyxl
from openpyxl import Workbook
import sys
from PySide2 import QtCore, QtGui, QtWidgets
from pathlib import Path
import os
import sqlite3
from sqlite3 import Error


##Fenetre utilisateur
class MaFenetre(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super().__init__()

        tabs = QtWidgets.QTabWidget()
        tabs.setTabPosition(QtWidgets.QTabWidget.North)
        tabs.setMovable(False)

        # les boutons

        self.boutonAchat = QtWidgets.QPushButton("Acquisto")
        self.boutonVente = QtWidgets.QPushButton("Vendita")

        self.boutonAddClient = QtWidgets.QPushButton("aggiungi cliente")
        self.boutonAddObjet = QtWidgets.QPushButton("Aggiungi articolo")
        self.bouttonSuppClient = QtWidgets.QPushButton("eliminare un cliente")

        # Les champs de texte
        self.__champTexte = QtWidgets.QLineEdit("")
        self.__champTexte.setPlaceholderText("Fattura1")
        self.labelMessage = QtWidgets.QLabel("")
        self.labelWarning = QtWidgets.QLabel("")

        self.labelAdd = QtWidgets.QLabel("")
        self.__champIva = QtWidgets.QLineEdit("")
        self.__champIva.setPlaceholderText("P.IVA")
        self.__champNom = QtWidgets.QLineEdit("")
        self.__champNom.setPlaceholderText("Nome della ditta")


        self.__champIIVA = QtWidgets.QLineEdit("")
        self.__champIIVA.setPlaceholderText("IVA del fornitore/cliente")
        self.__champCode = QtWidgets.QLineEdit("")
        self.__champCode.setPlaceholderText("Code")
        self.__champObjet = QtWidgets.QLineEdit("")
        self.__champObjet.setPlaceholderText("Desczione")
        self.__champQuantite = QtWidgets.QLineEdit("")
        self.__champQuantite.setPlaceholderText("Quantita")
        self.labelObjet = QtWidgets.QLabel("")

        layout1 = QtWidgets.QGridLayout()
        layout1.addWidget(self.__champTexte, 1, 1)
        layout1.addWidget(self.labelMessage, 2, 1)
        layout1.addWidget(self.boutonAchat, 3, 2)
        layout1.addWidget(self.boutonVente, 3, 0)
        layout1.addWidget(self.labelWarning,3,1)


        widget1 = QtWidgets.QWidget()
        widget1.setLayout(layout1)
        tabs.addTab(widget1, "Inserire una fattura")

        layout2 = QtWidgets.QGridLayout()

        layout2.addWidget(self.__champIva, 1, 0)
        layout2.addWidget(self.__champNom, 1, 2)
        layout2.addWidget(self.boutonAddClient, 3, 0)
        layout2.addWidget(self.labelAdd, 2, 1)
        layout2.addWidget(self.bouttonSuppClient, 3, 2)
        widget2 = QtWidgets.QWidget()
        widget2.setLayout(layout2)
        tabs.addTab(widget2, "Gestire i clienti")

        layout3 = QtWidgets.QGridLayout()
        layout3.addWidget(self.boutonAddObjet, 3, 4)
        layout3.addWidget(self.__champIIVA, 1, 1)
        layout3.addWidget(self.__champCode, 1, 2)
        layout3.addWidget(self.__champObjet, 1, 3)
        layout3.addWidget(self.__champQuantite, 1, 4)
        layout3.addWidget(self.labelObjet,2,2)
        widget3 = QtWidgets.QWidget()
        widget3.setLayout(layout3)
        tabs.addTab(widget3,"Gestione degli oggetti")

        self.setCentralWidget(tabs)
        #disposition widget fenetre



        icone = QtGui.QIcon()
        rep = os.getcwd()
        os.chdir(os.pardir)
        icone.addPixmap(QtGui.QPixmap("resources/bill.svg"))
        os.chdir(rep)
        self.setWindowIcon(icone)
        self.setWindowTitle("AutoFatture")

        self.boutonAchat.clicked.connect(self.genererAchat)
        self.boutonVente.clicked.connect(self.genererVente)
        self.boutonAddClient.clicked.connect(self.AddClientBouton)
        self.boutonAddObjet.clicked.connect(self.AddObjetBouton)
        self.bouttonSuppClient.clicked.connect(self.SuppClientBouton)

    ##Fonction appelé par le bouton Acquisto
    def genererAchat(self):

        path = Path("../Fatture_Acquisto/" + self.__champTexte.text() + ".pdf")
        rep = os.getcwd()
        # ouvre le pdf
        try:
            pdf = pdfplumber.open(path)
            os.chdir(rep)
        # cree le pdf
        except FileNotFoundError:
            self.__champTexte.clear()
            self.labelMessage.setText("Questa fattura non esiste")
            os.chdir(rep)
            return

        max_r1 = sheet1.max_row  # Donne l'emplacement pour écrire dans l'excel, dernière ligne remplie

        page1 = pdf.pages[0]
        # Recherche des infos importantes
        # Nom de la société
        Deno = page1.extract_tables()[0][0][0].find('Denominazione')
        Regime = page1.extract_tables()[0][0][0].find('Regime')
        name = page1.extract_tables()[0][0][0][Deno + 15:Regime - 1]

        # IVA De la société
        Iden = page1.extract_tables()[0][0][0].find('IVA:')
        Deno = page1.extract_tables()[0][0][0].find('Denominazione')
        IVA = page1.extract_tables()[0][0][0][Iden + 6:Deno - 1]
        if len(IVA) >= 12:
            IVA = IVA[1:13]
        IVA = IVA.split('\n')[0]
        # Date de la commande
        Date = page1.extract_tables()[1][1][3]

        # Numero de la commande
        NumCom = page1.extract_tables()[1][1][2]

        I = str(IVA)
        D = str(Date)
        N = str(NumCom)

        # On regarde si la facture a déjà été rentrée
        if checkA(I, D, N):
            self.labelMessage.setText("Questa fattura è già stata registrata")
            self.__champTexte.clear()
            wb.save('Fatture.xlsx')
            wb.close()
            return  # stop

        # On met le fournisseur si il n'existe pas
        if checkFourn(I):
            insert_fournisseur = '''INSERT INTO Fournisseurs(Nom,IVA)
                                        VALUES(?,?)'''
            list_fournisseurs = '''SELECT * 
                                    FROM Fournisseurs'''
            tuple = (name, I)
            cur.execute(insert_fournisseur, tuple)# on le met dans la bdd
            cur.execute(list_fournisseurs)
            pointeur = 1
            for ligne in cur:
                pointeur += 1
                for o in range(0,len(ligne)):
                    sheet4.cell(row=pointeur, column=2+o).value = ligne[o]

            conn.commit()

        # On ecrit les infos de la facture dans l'excel

        sheet1.cell(max_r1 + 2, 1).value = name
        sheet1.cell(max_r1 + 2, 2).value = Date

        # On rentre les produits dans l'excel d'achat et dans la bdd inventaire

        for page in pdf.pages:
            max_r1 = sheet1.max_row  # Donne l'emplacement pour écrire dans l'excel

            # Produits
            Lines = []
            infos = page.extract_tables(table_settings={"horizontal_strategy": "text"})
            for lines in infos:
                for line in lines:
                    if (len(line) > 7) and line[2] != '' and line[2] != 'Quantità':
                        Lines.append(line)
            for i in range(len(Lines)):
                code = Lines[i][0]
                desc = Lines[i][1]
                quant = Lines[i][2]
                if quant=='' or desc=='':
                    break
                if type(quant)==str:
                    try:
                        quant = float(quant.strip().split(" ")[0].replace(',', '.'))
                    except(ValueError):
                        return False
                for k in range(len(Lines[i])):
                    sheet1.cell(row=max_r1 + i + 2, column=k + 3).value = Lines[i][k]


                # il n'existe pas dans la BDD, on le rentre
                if checkObjet(code,desc):
                    insert_objet = '''INSERT INTO Inventaire(IVA,Code,Descrizione,Quantita)
                                    VALUES(?,?,?,?)'''
                    tuple_o = (IVA,code, desc, quant)
                    cur.execute(insert_objet, tuple_o)


                # Sinon on augmente sa quantité #a faire
                else:

                    nb_objets_request = '''SELECT Quantita FROM Inventaire
                                            WHERE Code = ? AND  IVA = ?'''
                    tuple_q = (code, IVA)
                    cur.execute(nb_objets_request, tuple_q)
                    rows = cur.fetchall()
                    quant_init=0
                    for row in rows:
                        quant_init = row[0]

                    if type(quant)==str:
                        quant = float(quant.strip().split(" ")[0].replace(',', '.'))
                    if type(quant_init)==str:
                        quant_init= float(quant_init.strip().split(" ")[0].replace(',', '.'))

                    tuple_o = (str(quant + quant_init), code, IVA)
                    update_objet = '''UPDATE Inventaire
                                        SET Quantita = ?
                                        WHERE Code = ? AND IVA = ?'''
                    cur.execute(update_objet, tuple_o)

        # On rentre la bdd inventaire dans la feuille 3

        check = """SELECT *
                FROM Inventaire as I"""
        cur.execute(check)
        pointeur = 1
        for ligne in cur:
            pointeur += 1
            for o in range(0, len(ligne)):
                sheet3.cell(row=pointeur, column=o + 2).value = ligne[o]

        # On rentre la facture dans la bdd facture

        insert_Fatture = '''INSERT INTO FattureA(IVA, Date,NumCom )
                            VALUES(?,?,?)'''
        tuple = (I, D, N)
        cur.execute(insert_Fatture, tuple)
        conn.commit()

        # On close tout
        wb.save('Fatture.xlsx')
        wb.close()
        self.labelMessage.setText("La fattura è stata aggiunta")
        self.__champTexte.clear()


    ##Fonction appelée par le bouton vendita
    def genererVente(self):

        path = Path("../Fatture_Vendita/" + self.__champTexte.text() + ".pdf")
        rep = os.getcwd()
        # ouvre le pdf
        try:
            pdf = pdfplumber.open(path)
            os.chdir(rep)
        # cree le pdf
        except FileNotFoundError:
            self.__champTexte.clear()
            self.labelMessage.setText("Questa fattura non esiste")
            os.chdir(rep)
            return



        max_r2 = sheet2.max_row  # Donne l'emplacement pour écrire dans l'excel

        # recherche des infos importantes
        page1 = pdf.pages[0]
        # date
        dataa = page1.extract_text().find('Data')
        sede = page1.extract_text().find('SEDE')
        Date = page1.extract_text()[dataa + 5: sede - 1]


        # entreprise
        dest = page1.extract_text().find('DESTINATARIO')
        copia = page1.extract_text().find('Copia')
        name = page1.extract_text()[dest + 13: copia - 1]

        # IVA
        IV = page1.extract_text().find('IVA')
        CF = page1.extract_text().find('C.F')
        IVA = page1.extract_text()[IV + 4: CF - 3]

        # NUM
        Nume = page1.extract_text().find('Numero')
        Data = page1.extract_text().find('Data')
        NumCom = page1.extract_text()[Nume + 7: Data - 1]

        I = str(IVA)
        D = str(Date)
        N = str(NumCom)
        # regarde si elle n'a pas déjà été rentrée

        if checkV(I, D, N):
            self.labelMessage.setText("Questa fattura è già stata registrata")
            self.__champTexte.clear()
            return
        # rentre le fournisseur

        if checkFourn(I):
            insert_fournisseur = '''INSERT INTO Fournisseurs(Nom,IVA)
                                        VALUES(?,?)'''
            list2_fournisseurs = '''SELECT * 
                                    FROM Fournisseurs'''
            tuple = (name, I)
            cur.execute(insert_fournisseur, tuple)  # on le met dans la bdd
            cur.execute(list2_fournisseurs)
            pointeur = 1
            for ligne in cur:
                pointeur += 1
                for o in range(0, len(ligne)):
                    sheet4.cell(row=pointeur, column=2 + o).value = ligne[o]

            conn.commit()

        sheet2.cell(max_r2 + 2, 1).value = name
        sheet2.cell(max_r2 + 2, 2).value = Date

        # Produits
        max_r2 = sheet2.max_row  # Donne l'emplacement pour écrire dans l'excel
        Lines = []
        for page in pdf.pages:
            table = page.extract_tables(table_settings={"vertical_strategy": "text"})
            for tables in table:
                for line in tables:
                    if not 'ALIQUOTE' in line[1]:
                        if not'ALIQUOTE' in line[2]:
                            Lines.append(line)
        for i in range(len(Lines)):
            code = Lines[i][1]
            sheet2.cell(row=max_r2 + i + 2, column=1+2).value = code
            #parsing selon les unités :
            for k in range(len(Lines[i])):
                if 'N' in Lines[i][k]:
                    o=Lines[i][k].split(' ')
                    if len(o)==1:
                        desc=[]
                        for _ in range(2,k):
                            desc.append(Lines[i][_])
                        desc=''.join(desc)
                        sheet2.cell(row=max_r2 + i + 2, column=2+2).value = desc
                        quant=Lines[i][k+1].split('\n')[0]
                        quant=- float(quant.strip().split(" ")[0].replace(',', '.'))
                        sheet2.cell(row=max_r2 + i + 2, column=6).value = - quant

                    else:
                        for l in range(len(o)):
                            if o[l]=='N':
                                q=o[:l]
                                if k>1:
                                    for _ in range(2,k):
                                        q.append(Lines[i][_])
                                desc=''.join(q)
                                sheet2.cell(row=max_r2 + i + 2, column=2+2).value = desc
                                #desc=Lines[i][k-1]+o[:l]
                                #desc=Lines[i][k-1]
                                quant=o[l+1].split('\n')[0]
                                quant=- float(quant.strip().split(" ")[0].replace(',', '.'))
                                sheet2.cell(row=max_r2 + i + 2, column=6).value = - quant

                    indice = k + 1
                    for q in range(k+1,len(Lines[i])):

                        if Lines[i][q]!='' or Lines[i][q]!=' ':
                            print(Lines[i][q])
                            sheet2.cell(row=max_r2 + i + 2, column=indice + 1).value = Lines[i][q]
                            indice+=1

                


            # Si il n'existe pas dans la BDD, message d'avertissement puis rentre
            if checkObjet(code,desc):

                self.labelWarning.setText("Si prega di notare che uno degli articoli venduti non esiste nel database")
                insert_objet='''INSERT INTO Inventaire(IVA,Code,Descrizione,Quantita)
                                        VALUES(?,?,?,?)'''
                tuple_o=(IVA,code,desc,quant)
                cur.execute(insert_objet, tuple_o)
            #Sinon, on change sa quantité
            else:

                nb_objets_request = '''SELECT Quantita FROM Inventaire
                                                        WHERE Code = ? AND  IVA = ?'''
                tuple_q = (code, IVA)
                cur.execute(nb_objets_request, tuple_q)
                rows = cur.fetchall()
                quant_init=0
                for row in rows:
                    quant_init = row[0]

                if type(quant) == str:
                    quant = float(quant.strip().split(" ")[0].replace(',', '.'))
                if type(quant_init) == str:
                    quant_init = float(quant_init.strip().split(" ")[0].replace(',', '.'))

                tuple_o = (str(quant + quant_init), code, IVA)
                update_objet = '''UPDATE Inventaire
                                                    SET Quantita = ?
                                                    WHERE Code = ? AND IVA = ?'''
                cur.execute(update_objet, tuple_o)

        # On rentre la bdd inventaire dans la feuille 3

        check = """SELECT *
                    FROM Inventaire as I"""
        cur.execute(check)
        pointeur = 1
        for ligne in cur:
            pointeur += 1
            for o in range(0, len(ligne)):
                sheet3.cell(row=pointeur, column=o + 2).value = ligne[o]

        # rentre la facture

        insert_Fatture = '''INSERT INTO FattureV(IVA, Date,NumCom )
                                    VALUES(?,?,?)'''
        tuple = (I, D, N)
        cur.execute(insert_Fatture, tuple)
        conn.commit()
        # fermeture
        wb.save('Fatture.xlsx')
        wb.close()
        self.labelMessage.setText("La fattura è stata aggiunta")
        self.__champTexte.clear()


    ## fonction appelé par le bouton add client
    def AddClientBouton(self):
        IVA = self.__champIva.text()
        Nom = self.__champNom.text()
        if len(IVA) > 11:
            self.labelAdd.setText("IVA troppo a lungo")
            self.__champIva.clear()
            return
        elif len(IVA) < 11:
            self.labelAdd.setText("IVA troppo corto")
            self.__champIva.clear()
            return
        else:

            if addClient(IVA, Nom):
                self.labelAdd.setText("Il cliente è stato aggiunto")
            else:
                self.labelAdd.setText("Il cliente esiste già")

            self.__champIva.clear()
            self.__champNom.clear()

    def SuppClientBouton(self):
        IVA = self.__champIva.text()
        if len(IVA) > 11:
            self.labelAdd.setText("IVA troppo lunga")
            self.__champIva.clear()
            return
        elif len(IVA) < 11:
            self.labelAdd.setText("IVA too short")
            self.__champIva.clear()
            return
        else:

            if SuppClient(IVA):
                self.labelAdd.setText("Il cliente è stato ritirato con successo")
            else:
                self.labelAdd.setText("Il cliente non esiste")

            self.__champIva.clear()
            self.__champNom.clear()

    ## Ce que fais le bouton Add Objet
    def AddObjetBouton(self):
        Code = self.__champCode.text()
        Objet = self.__champObjet.text()
        Quantite = self.__champQuantite.text()
        IVA = self.__champIIVA.text()
        if len(IVA) > 11:
            self.labelObjet.setText("IVA troppo a lungo")
            return
        elif len(IVA) < 11:
            self.labelObjet.setText("IVA troppo corto")
            return
        if addObjet(IVA,Code,Objet,Quantite):
            self.labelObjet.setText("Il prodotto è stato aggiunto")
        else:
            self.labelObjet.setText("Errore di quantità")

        self.__champCode.clear()
        self.__champObjet.clear()
        self.__champQuantite.clear()
        self.__champIIVA.clear()



## Verifie si la facture achat n'existe pas déjà dans la bdd
def checkA(IVA, Date, NumCom):
    tuple = (str(IVA), str(Date), str(NumCom))
    check = """SELECT IVA,Date, NumCom
    FROM FattureA as F
    WHERE F.IVA = ? AND F.Date = ? AND F.NumCom = ?"""
    cur.execute(check, tuple)
    boo = ''
    for i in cur:
        boo = i
    if (boo == ''):
        return False
    return True


## Verifie si la facture vente n'existe pas déjà dans la bdd
def checkV(IVA, Date, NumCom):
    tuple = (str(IVA), str(Date), str(NumCom))
    check = """SELECT IVA,Date, NumCom
    FROM FattureV as F
    WHERE F.IVA = ? AND F.Date = ? AND F.NumCom = ?"""
    cur.execute(check, tuple)
    boo = ''
    for i in cur:
        boo = i
    if (boo == ''):
        return False
    return True


## Regarde si le fournisseur existe déjà
def checkFourn(IVA):
    tuple = (str(IVA),)
    check = """SELECT IVA
        FROM Fournisseurs as F
        WHERE F.IVA= ?  """
    cur.execute(check, tuple)
    bool = ''
    for row in cur:
        bool = row
    if bool == '':
        return True
    return False


## Ajoute le client à la main
def addClient(IVA, Nom):
    if (checkFourn(IVA)):
        tuple = (str(IVA), str(Nom))
        requestAdd = """INSERT INTO Fournisseurs(IVA,Nom)
                            VALUES(?,?)"""
        list_fournisseurs4 = '''SELECT * 
                                            FROM Fournisseurs'''
        cur.execute(requestAdd, tuple)
        cur.execute(list_fournisseurs4)
        pointeur = 1

        for ligne in cur:
            pointeur += 1
            for o in range(0, len(ligne)):
                sheet4.cell(row=pointeur, column=2 + o).value = ligne[o]

        conn.commit()
        wb.save('Fatture.xlsx')
        return True
    else:
        return False

## Ajoute un objet à la main
def addObjet(IVA,Code,Objet,Quantite):
    if checkObjet(Code,Objet): #l'objet n'existe pas, on le rentre
        try:
            Quantite = float(Quantite.strip().split(" ")[0].replace(',', '.'))

        except(ValueError):
            return False
        tuple=(IVA,str(Code), str(Objet),str(Quantite))
        request = """INSERT INTO Inventaire(IVA,Code, Descrizione, Quantita)
                        VALUES(?,?,?,?)"""
        cur.execute(request, tuple)
        conn.commit()
        return True
    else: #l'objet existe déjà : calcul time
        tuple=(str(Code),)
        check = '''SELECT Quantita FROM Inventaire as I
                            WHERE I.Code = ?'''
        cur.execute(check, tuple)
        QInit=[line for line in cur][0][0]
        if type(QInit)==str:
            try:
                QInitFloat = float(QInit.strip().split(" ")[0].replace(',','.'))
            except(ValueError):
                return False
        else:
            QInitFloat=QInit
        if type(Quantite)==int or type(Quantite)==float:
            QFinal = Quantite+QInitFloat
        else:
            try:
                Quantite= float(Quantite.strip().split(" ")[0].replace(',','.'))
            except(ValueError):
                return False
            QFinal = Quantite + QInitFloat
        check2="""UPDATE Inventaire 
                    SET Quantita = ?
                    WHERE Code=?"""
        tuple=(QFinal,str(Code))
        cur.execute(check2, tuple)
        conn.commit()
        return True


## Possiblement pour supprimer un client
def SuppClient(IVA):
    if (checkFourn(IVA)):
        return False
    else:
        tuple = (str(IVA),)
        requestSupp = """DELETE FROM Fournisseurs
                            WHERE IVA=?"""
        cur.execute(requestSupp, tuple)
        conn.commit()
        return True

## Est vrai si l'objet n'existe pas dans la BDD
def checkObjet(code,desc):
    tuple = (code,desc)
    check = '''SELECT * FROM Inventaire as I
                    WHERE (I.Code = ? OR I.Descrizione = ?)'''
    cur.execute(check,tuple)
    b=''
    for row in cur:
        b=row
    if b=='':
        return True
    return False



### Code principal

#creation excel
try:
    wb = openpyxl.load_workbook('Fatture.xlsx')
    sheet1 = wb['Acquista']
    sheet2 = wb['Vendita']
    sheet3 = wb['Inventario']
    sheet4 = wb['Fornitori']

except FileNotFoundError:  # Sinon, on le crée
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Acquista'
    wb.create_sheet('Vendita')
    wb.create_sheet('Inventario')
    wb.create_sheet('Fornitori')
    sheet1 = wb['Acquista']
    sheet2 = wb['Vendita']
    sheet3 = wb['Inventario']
    sheet4 = wb['Fornitori']
    sheet1.cell(1, 3).value = 'Cod Articolo'
    sheet1.cell(1, 4).value = 'Desczione'
    sheet1.cell(1, 5).value = 'Quantita'
    sheet1.cell(1, 6).value = 'Prezzo unitario'
    sheet1.cell(1, 7).value = 'UM'
    sheet1.cell(1, 8).value = 'Sconto o magg'
    sheet1.cell(1, 9).value = '%IVA'
    sheet1.cell(1, 10).value = 'Prezzo totale'
    sheet2.cell(1, 3).value = 'ARTICOLO'
    sheet2.cell(1, 4).value = 'DESCRIZIONE'
    sheet2.cell(1, 5).value = 'UNITÀ'
    sheet2.cell(1, 6).value = 'Q.TÀ'
    sheet2.cell(1, 7).value = 'IMPORTO U.'
    sheet2.cell(1, 8).value = 'IVA%SCONTO'
    sheet2.cell(1, 9).value = 'IMPORTO'
    sheet3.cell(1, 3).value = 'Acquista'
    sheet3.cell(1, 12).value = 'Vendita'
    sheet4.cell(1, 1).value = 'Fornitori'
    sheet3.cell(1, 1).value = 'Inventario'
    wb.save('Fatture.xlsx')


#creation bdd
conn = None
rep = os.getcwd()
os.chdir(os.pardir)
dbf = str(Path("DB/database.db").absolute())
os.chdir(rep)
# cree la bdd
try:
    conn = sqlite3.connect(dbf)

    cur = conn.cursor()
    table_facturesA = '''CREATE TABLE IF NOT EXISTS FattureA(
                              IVA TEXT,
                              Date TEXT,
                              NumCom TEXT
                             )'''
    cur.execute(table_facturesA)
    table_facturesV = '''CREATE TABLE IF NOT EXISTS FattureV(
                                  IVA TEXT,
                                  Date TEXT,
                                  NumCom TEXT
                                 )'''
    cur.execute(table_facturesV)
    table_fournisseurs = '''CREATE TABLE IF NOT EXISTS Fournisseurs(
                                Nom TEXT,
                                IVA TEXT
                                )'''
    cur.execute(table_fournisseurs)
    table_inventaire = '''CREATE TABLE IF NOT EXISTS Inventaire(
                                IVA TEXT,
                                Code TEXT,
                                Descrizione TEXT,
                                Quantita TEXT
                                )'''

    cur.execute(table_inventaire)
    conn.commit()
    #appel la classe fenetre
    app = QtWidgets.QApplication(sys.argv)

    window = MaFenetre()
    window.show()

    app.exec_()

except Error as e:
    print(e)
finally:
    if conn:
       conn.close()
