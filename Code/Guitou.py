import pdfplumber
import openpyxl
from openpyxl import Workbook
import sys
from PySide2 import QtCore, QtGui, QtWidgets


class MaFenetre(QtWidgets.QDialog):
    def __init__(self, parent=None):
        QtWidgets.QDialog.__init__(self, parent)

        self.__boutonGenerer = QtWidgets.QPushButton("Générer")
        # Le champ de texte
        self.__champTexte = QtWidgets.QLineEdit("")

        layout = QtWidgets.QGridLayout()
        layout.addWidget(self.__champTexte, 2, 1)

        layout.addWidget(self.__boutonGenerer, 3, 2)
        self.setLayout(layout)

        icone = QtGui.QIcon()
        icone.addPixmap(QtGui.QPixmap("cadenas.svg"))
        self.setWindowIcon(icone)

        self.__boutonGenerer.clicked.connect(self.generer)

    def generer(self):

        path = self.__champTexte.text()

        pdf = pdfplumber.open(path)
        # print(_)
        try:  # Test si l'excel existe, dans ce cas là, on l'ouvre
            wb = openpyxl.load_workbook('test.xlsx')
            sheet1 = wb.active

        except FileNotFoundError:  # Sinon, on le crée
            wb = Workbook()
            sheet1 = wb.active
            sheet1.cell(1, 3).value = 'Cod Articolo'
            sheet1.cell(1, 4).value = 'Desczione'
            sheet1.cell(1, 5).value = 'Quantita'
            sheet1.cell(1, 6).value = 'Prezzo unitario'
            sheet1.cell(1, 7).value = 'UM'
            sheet1.cell(1, 8).value = 'Sconto o magg'
            sheet1.cell(1, 9).value = '%IVA'
            sheet1.cell(1, 10).value = 'Prezzo totale'

        max_r = sheet1.max_row  # Donne l'emplacement pour écrire dans l'excel

        page1 = pdf.pages[0]
        # Nom de la société
        Deno = page1.extract_tables()[0][0][0].find('Denominazione')
        Regime = page1.extract_tables()[0][0][0].find('Regime')
        name = page1.extract_tables()[0][0][0][Deno + 15:Regime]
        sheet1.cell(max_r + 2, 1).value = name
        # Date
        date = page1.extract_tables()[1][1][3]
        sheet1.cell(max_r + 2, 2).value = date

        for page in pdf.pages:
            max_r = sheet1.max_row  # Donne l'emplacement pour écrire dans l'excel

            # Produits
            Lines = []
            x = page.extract_tables(table_settings={"horizontal_strategy": "text"})
            for lines in x:
                for line in lines:
                    if (len(line) > 7) and line[2] != '' and line[2] != 'Quantità':
                        Lines.append(line)
                        # print(line)
                        # """x = table[1] #a ameliorer
                    for i in range(len(Lines)):
                        # print(Lines[i])
                        # s = line[i]
                        # lis = s.split("\n")
                        for k in range(len(Lines[i])):
                            sheet1.cell(row=max_r + i + 2, column=k + 3).value = Lines[i][k]
                        # if (s != ''):
                        # print(s)

        wb.save('test.xlsx')

        wb.close()
# """


if __name__ == '__main__':




    app = QtWidgets.QApplication(sys.argv)
    dialog = MaFenetre()
    dialog.exec_()
