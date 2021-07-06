import pdfplumber
import openpyxl
from openpyxl import Workbook

if __name__ == '__main__':
    wb = openpyxl.load_workbook('Fatture.xlsx')
    sheet1 = wb['Acquista']
    sheet2 = wb['Vendita']
    sheet3 = wb['Inventario']
    sheet4 = wb['Fornitori']
    #rentrée fournisseur
    list_fournisseurs = '''SELECT * 
                                        FROM Fournisseurs'''
    cur.execute(list_fournisseurs)
    pointeur = 1
    for ligne in cur:
        pointeur += 1
        for o in range(0, len(ligne)):
            sheet4.cell(row=pointeur, column=2 + o).value = ligne[o]
    #rentrée inventaire
    check = """SELECT *
                    FROM Inventaire as I"""
    cur.execute(check)
    pointeur = 1
    for ligne in cur:
        pointeur += 1
        for o in range(0, len(ligne)):
            sheet3.cell(row=pointeur, column=o + 2).value = ligne[o]

    wb.save('Fatture.xlsx')
