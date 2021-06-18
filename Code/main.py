import pdfplumber
import openpyxl
from openpyxl import Workbook

if __name__ == '__main__':

    #for _ in range(1,13):
    #path = 'Facture'+str(_)+'.pdf'
    path = 'Facture1.pdf'

    pdf = pdfplumber.open(path)
    #print(_)
    try: #Test si l'excel existe, dans ce cas là, on l'ouvre
        wb = openpyxl.load_workbook('test.xlsx')
        sheet1 = wb.active

    except FileNotFoundError: #Sinon, on le crée
        wb = Workbook()
        sheet1 = wb.active
        sheet1.cell(1, 3).value='Cod Articolo'
        sheet1.cell(1, 4).value = 'Desczione'
        sheet1.cell(1, 5).value = 'Quantita'
        sheet1.cell(1, 6).value = 'Prezzo unitario'
        sheet1.cell(1, 7).value = 'UM'
        sheet1.cell(1, 8).value = 'Sconto o magg'
        sheet1.cell(1, 9).value = '%IVA'
        sheet1.cell(1, 10).value = 'Prezzo totale'

    max_r = sheet1.max_row #Donne l'emplacement pour écrire dans l'excel

    page1=pdf.pages[0]
    # Nom de la société
    Deno = page1.extract_tables()[0][0][0].find('Denominazione')
    Regime = page1.extract_tables()[0][0][0].find('Regime')
    name = page1.extract_tables()[0][0][0][Deno + 15:Regime]
    sheet1.cell(max_r + 2, 1).value = name
    # Date
    date = page1.extract_tables()[1][1][3]
    sheet1.cell(max_r + 2, 2).value = date



    for page in pdf.pages:
        #Produits
        for k in range (len(page.extract_tables())): #parcours des tables
            header=page.extract_tables()[k][0][0] #cherche si le tableau est un tableau avec des articles
            table=page.extract_tables()[k]
            #if 'articolo' in header:
            print(table)

            """x = table[1] #a ameliorer
                for i in range(len(x)):
                    s = x[i]
                    lis = s.split("\n")
                    for k in range(len(lis)):
                        sheet1.cell(row=max_r+k + 2, column=i + 3).value=lis[k]
                    if (s != ''):
                        print(s)




    wb.save('test.xlsx')


    wb.close()
#"""
