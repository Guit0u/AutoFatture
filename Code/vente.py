import pdfplumber
import openpyxl
from openpyxl import Workbook

if __name__ == '__main__':


    path = 'Fattura_27.PDF'
    pdf = pdfplumber.open(path)


    try:  # Test si l'excel existe, dans ce cas là, on l'ouvre
        wb = openpyxl.load_workbook('test.xlsx')
        sheet2=wb['Vendita']
        sheet1 = wb['Acquista']
        sheet3 = wb['Inventario']


    except FileNotFoundError:  # Sinon, on le crée
        wb = Workbook()
        sheet=wb.active
        sheet.title='Acquista'
        wb.create_sheet('Vendita')
        wb.create_sheet('Inventario')
        sheet1 = wb['Acquista']
        sheet2 = wb['Vendita']
        sheet3 = wb['Inventario']
        sheet2.cell(1, 3).value = 'ARTICOLO'
        sheet2.cell(1, 4).value = 'DESCRIZIONE'
        sheet2.cell(1, 5).value = 'UNITÀ'
        sheet2.cell(1, 6).value = 'Q.TÀ'
        sheet2.cell(1, 7).value = 'IMPORTO U.'
        sheet2.cell(1, 8).value = 'IVA%SCONTO'
        sheet2.cell(1, 9).value = 'IMPORTO'
        sheet1.cell(1, 3).value = 'Cod Articolo'
        sheet1.cell(1, 4).value = 'Desczione'
        sheet1.cell(1, 5).value = 'Quantita'
        sheet1.cell(1, 6).value = 'Prezzo unitario'
        sheet1.cell(1, 7).value = 'UM'
        sheet1.cell(1, 8).value = 'Sconto o magg'
        sheet1.cell(1, 9).value = '%IVA'
        sheet1.cell(1, 10).value = 'Prezzo totale'
        sheet3.cell(1, 3).value = 'Acquista'
        sheet3.cell(1, 12).value = 'Vendita'

    max_r = sheet2.max_row  # Donne l'emplacement pour écrire dans l'excel

    page1 = pdf.pages[0]
    # date
    x = page1.extract_text().find('Data')
    y = page1.extract_text().find('SEDE')
    date = page1.extract_text()[x + 5: y]
    sheet2.cell(max_r + 2, 2).value = date
    # entreprise
    z = page1.extract_text().find('DESTINATARIO')
    t = page1.extract_text().find('Copia')
    name = page1.extract_text()[z + 12: t]
    sheet2.cell(max_r + 2, 1).value = name



    fichier = open("data.txt", "a")
    fichier.write(name)
    fichier.write(date)
    fichier.close()

    #Produits
    max_r = sheet2.max_row  # Donne l'emplacement pour écrire dans l'excel
    Lines = []
    for page in pdf.pages:
        table = page.extract_tables(table_settings={"vertical_strategy": "text"})
        for tables in table:
            for line in tables:
                if not 'ALIQUOTE' in line[1]:
                    Lines.append(line)
                for i in range(len(Lines)):
                    #print(Lines[i])
                    #s = line[i]
                    #lis = s.split("\n")
                    for k in range(len(Lines[i])):
                        sheet2.cell(row=max_r + i + 2, column=k + 2).value = Lines[i][k]
                        sheet3.cell(row=max_r + i +1, column=k + 11).value = Lines[i][k]
                        #print(c)


    wb.save('test.xlsx')

    wb.close()
